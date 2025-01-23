import os
import sys
import asyncio
import tempfile
import traceback
import logging
import streamlit as st
from pathlib import Path
from collections import deque
from openpyxl import load_workbook
from googletrans import Translator
from datetime import datetime, timezone

def configure_logger(log_level='INFO', log_file_path=None):
  """
  Configures a Python logger with the specified log level and optional file output.

  Args:
    log_level: The logging level (e.g., 'DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'). 
                Defaults to 'INFO'.
    log_file_path: Optional path to the log file. If None, logs to console only.

  Returns:
    The configured logger object.
  """

  logger = logging.getLogger(__name__)
  logger.setLevel(log_level)

  # Default message format
  formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

  # Create console handler
  console_handler = logging.StreamHandler()
  console_handler.setFormatter(formatter)
  logger.addHandler(console_handler)

  # Create file handler if log_file_path is provided
  if log_file_path:
    file_handler = logging.FileHandler(log_file_path)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

  return logger


date = datetime.now(timezone.utc)
date_string = f"{date.year}-{date.month}-{date.day}"
log_file = f'{date_string}-translate_app.log'
logger = configure_logger(log_level='DEBUG', log_file_path=log_file) 

########## row_ds data structure ##############
class CancellationException(Exception):
    pass

class CancellationToken:
    def __init__(self):
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def is_cancelled(self):
        return self._is_cancelled


class TranslateRow:

    def __init__(self, row):
        self.row = row
        self.row_len = len(self.row)
        self.no_translate_queue = deque()
        self.no_translate_idx = []
        self.pre_translate_queue = deque()
        self.post_translate_queue = deque()
        self.rebuilt_queue = deque()

    @staticmethod
    def no_translate_cell(cell):

        if not isinstance(cell.value, str):
            return True
        
        if not cell.value:
            return True
        
        if cell.value.startswith("="):
            return True
        
        return False

    def prepare_data_to_translate(self):


        for idx, cell in enumerate(self.row):
            if self.no_translate_cell(cell):
                self.no_translate_idx.append(idx)
                self.no_translate_queue.append(cell.value)
            else:
                self.pre_translate_queue.append(cell.value)


    async def translate_row(self, src, dest, cancellation_token):
        async with Translator() as translator:
            if cancellation_token.is_cancelled():
                raise CancellationException
            translations = await translator.translate(list(self.pre_translate_queue), src=src, dest=dest)

            for translation in translations:
                self.post_translate_queue.append(translation.text)


    def post_translation_rebuild(self):

        assert len(self.pre_translate_queue) == len(self.post_translate_queue), "Pre and post translate queue are of different length"

        for idx in range(self.row_len):
            if idx in self.no_translate_idx:
                self.rebuilt_queue.append(self.no_translate_queue.popleft())
            else:
                self.rebuilt_queue.append(self.post_translate_queue.popleft())


    async def perform_translation(self, src, dest, cancellation_token):

        self.prepare_data_to_translate()

        try:
            await self.translate_row(src, dest, cancellation_token)
            self.post_translation_rebuild()
            return list(self.rebuilt_queue)
        except Exception as exc:
            t = traceback.format_exc()
            logger.error(f"Error originates: {t}")
            logger.error(f"Error while translating {exc}")


######## Handlers file content ########

async def translate_sheet(ws, src_lang, dest_lang, cancellation_token, log_queue):
    """Translates all cells in a sheet asynchronously."""
    ws_title = ws.title
    logger.info(f"Started translation of sheet {ws_title}")
    tasks = []
    for row in ws.iter_rows():
        t_row = TranslateRow(row)
        tasks.append(asyncio.create_task(t_row.perform_translation(src_lang, dest_lang, cancellation_token)))

    translated_values = await asyncio.gather(*tasks)

    for row, row_values in zip(ws.iter_rows(), translated_values):
        for cell, value in zip(row, row_values):
            if value is not None:  # Skip cancelled translations
                cell.value = value

    logger.info(f"Ended translation of sheet {ws_title}")
    await log_queue.put(f"Worksheet '{ws_title}' translated.")

async def translate_workbook(input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=False):
    """Translates an Excel workbook asynchronously."""
    wb = load_workbook(input_file, keep_vba=True)  # Macros are preserved
    output_file = os.path.join(os.path.dirname(input_file), f"translated_{os.path.basename(input_file)}")

    tasks = []
    total_sheets = len(wb.sheetnames)
    progress = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tasks.append(asyncio.create_task(translate_sheet(ws, src_lang, dest_lang, cancellation_token, log_queue)))
        progress += 1
        progress_bar.progress(progress / total_sheets)

    await asyncio.gather(*tasks)

    if op_in_dir:
        file_dir = os.path.dirname(input_file)
        new_dir = os.path.join(file_dir, f"translated_files")
        output_file = os.path.join(os.path.dirname(new_dir), f"translated_{os.path.basename(input_file)}")

    wb.save(output_file)
    return output_file

async def translate_file(input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=False):
    """Translates a single Excel file asynchronously."""
    try:
        translated_file_path = await translate_workbook(input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=op_in_dir)
        await log_queue.put(f"File translated successfully: {translated_file_path}")
        return translated_file_path
    except CancellationException:
        await log_queue.put("File translation cancelled.")
    except Exception as e:
        logger.error(f"Error translating file: {e}")
        raise RuntimeError(f"An error occurred while translating the file: {e}")

async def translate_folder(folder_path, src_lang, dest_lang, cancellation_token, log_queue, progress_bar):
    """Translates all Excel files within a specified folder asynchronously."""
    excel_files = [file for file in Path(folder_path).rglob('*.xlsx') or Path(folder_path).rglob('*.xlsm')]
    logger.debug(f"files retrieved: {excel_files}")
    tasks = []
    file_paths = []
    total_files = len(excel_files)
    progress = 0

    for file_path in excel_files:
        logger.info(f"Started translation for workbook at: {file_path}")
        tasks.append(asyncio.create_task(translate_file(str(file_path), src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=True)))
        file_paths.append(file_path)
        progress += 1
        progress_bar.progress(progress / total_files)
        logger.info(f"Completed translation for workbook at: {file_path}")
    results = await asyncio.gather(*tasks)

    for file_path, result in zip(file_paths, results):
        if result:
            await log_queue.put(f"Translated: {file_path} -> {result}")


##### start of streamlit app ##########
# Ensure the script can access the utils folder
if getattr(sys, 'frozen', False):
    # If the app is packaged, get the path of the executable
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # If running normally, get the current directory
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

sys.path.append(os.path.join(BASE_DIR, 'utils'))
TEMP_DIR = os.path.join(BASE_DIR, "temp")
os.makedirs(TEMP_DIR, exist_ok=True)

def main():
    

    # Disable the Streamlit hamburger menu and footer for a cleaner look
    st.markdown(
        """
        <style>
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title("Excel Translator (macOS-Friendly)")
    st.write("Translate Excel files while streaming logs.")

    option = st.radio("Select Option", ("Single File", "Folder"))

    src_language_options = {
        "auto": "Auto Detect",
        "de": "German",
        "en": "English",
        "fr": "French",
        "es": "Spanish",
        "it": "Italian",
        "pt": "Portuguese",
        "ja": "Japanese",
        "zh-CN": "Chinese (Simplified)",
        "ru": "Russian"
    }

    desc_language_options = {
        "en": "English",
        "de": "German",
        "fr": "French",
        "es": "Spanish",
        "it": "Italian",
        "pt": "Portuguese",
        "ja": "Japanese",
        "zh-CN": "Chinese (Simplified)",
        "ru": "Russian"
    }

    src_lang = st.selectbox("Source Language", list(src_language_options.keys()), format_func=lambda x: src_language_options[x])
    dest_lang = st.selectbox("Destination Language", list(desc_language_options.keys()), format_func=lambda x: desc_language_options[x])

    log_queue = asyncio.Queue()

    if option == "Single File":
        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsm", "xlsx"])

        if uploaded_file:
            st.write("Uploaded file:", uploaded_file.name)

            if st.button("Translate File"):
                try:
                    temp_input_file = os.path.join(TEMP_DIR, uploaded_file.name)
                    with open(temp_input_file, "wb") as f:
                        f.write(uploaded_file.read())

                    cancellation_token = CancellationToken()
                    progress_bar = st.progress(0)

                    # Run the translation synchronously
                    translated_file_path = asyncio.run(
                        translate_file(temp_input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar)
                    )

                    # Display logs
                    while not log_queue.empty():
                        log_message = log_queue.get_nowait()
                        st.info(log_message)

                    if translated_file_path:
                        st.success("Translation completed. Download your file below:")
                        with open(translated_file_path, "rb") as f:
                            st.download_button(
                                "Download Translated File",
                                f,
                                file_name=os.path.basename(translated_file_path),
                            )

                except Exception as e:
                    st.error(f"An error occurred: {e}")

    else:
        
        folder_path = st.text_input("Enter folder path containing Excel files")

        if st.button("Translate Folder"):
            try:
                cancellation_token = CancellationToken()
                progress_bar = st.progress(0)

                async def process_folder():
                    await translate_folder(folder_path, src_lang, dest_lang, cancellation_token, log_queue, progress_bar)
                    st.success("Translation of all files in the folder completed.")

                asyncio.run(process_folder())

                while not log_queue.empty():
                    log_message = log_queue.get_nowait()
                    st.info(log_message)

            except Exception as e:
                st.error(f"An error occurred: {e}")

if __name__ == "__main__":
    import streamlit as st  # Import streamlit within the main block
    st.set_page_config(page_title="Excel Translator", layout="centered")

    main()  # Call your main function as before

    # Launch the Streamlit app
    st.write("Streamlit app started.") 
    st.balloons()  # Add a visual cue