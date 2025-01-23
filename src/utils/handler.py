import os
import asyncio
from pathlib import Path
from openpyxl import load_workbook

from utils.logging_mech import logger as logging
from utils.row_ds import TranslateRow, CancellationException

async def translate_sheet(ws, src_lang, dest_lang, cancellation_token, log_queue):
    """Translates all cells in a sheet asynchronously."""
    ws_title = ws.title
    logging.info(f"Started translation of sheet {ws_title}")
    tasks = []
    for row in ws.iter_rows():
        t_row = TranslateRow(row)
        tasks.append(asyncio.create_task(t_row.perform_translation(src_lang, dest_lang, cancellation_token)))

    translated_values = await asyncio.gather(*tasks)

    for row, row_values in zip(ws.iter_rows(), translated_values):
        for cell, value in zip(row, row_values):
            if value is not None:  # Skip cancelled translations
                cell.value = value

    logging.info(f"Ended translation of sheet {ws_title}")
    await log_queue.put(f"Worksheet '{ws_title}' translated.")

async def translate_workbook(input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=False):
    """Translates an Excel workbook asynchronously."""
    wb = load_workbook(input_file, keep_vba=True)  # Macros are preserved
    output_file = os.path.join(os.path.dirname(input_file), f"translated_{os.path.basename(input_file)}")

    tasks = []
    total_sheets = len(wb.sheetnames)
    progress = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tasks.append(asyncio.create_task(translate_sheet(ws, src_lang, dest_lang, cancellation_token, log_queue)))
        progress += 1
        progress_bar.progress(progress / total_sheets)

    await asyncio.gather(*tasks)

    if op_in_dir:
        file_dir = os.path.dirname(input_file)
        new_dir = os.path.join(file_dir, f"translated_files")
        output_file = os.path.join(os.path.dirname(new_dir), f"translated_{os.path.basename(input_file)}")

    wb.save(output_file)
    return output_file

async def translate_file(input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=False):
    """Translates a single Excel file asynchronously."""
    try:
        translated_file_path = await translate_workbook(input_file, src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=op_in_dir)
        await log_queue.put(f"File translated successfully: {translated_file_path}")
        return translated_file_path
    except CancellationException:
        await log_queue.put("File translation cancelled.")
    except Exception as e:
        logging.error(f"Error translating file: {e}")
        raise RuntimeError(f"An error occurred while translating the file: {e}")

async def translate_folder(folder_path, src_lang, dest_lang, cancellation_token, log_queue, progress_bar):
    """Translates all Excel files within a specified folder asynchronously."""
    excel_files = [file for file in Path(folder_path).rglob('*.xlsx') or Path(folder_path).rglob('*.xlsm')]
    logging.debug(f"files retrieved: {excel_files}")
    tasks = []
    file_paths = []
    total_files = len(excel_files)
    progress = 0

    for file_path in excel_files:
        logging.info(f"Started translation for workbook at: {file_path}")
        tasks.append(asyncio.create_task(translate_file(str(file_path), src_lang, dest_lang, cancellation_token, log_queue, progress_bar, op_in_dir=True)))
        file_paths.append(file_path)
        progress += 1
        progress_bar.progress(progress / total_files)
        logging.info(f"Completed translation for workbook at: {file_path}")
    results = await asyncio.gather(*tasks)

    for file_path, result in zip(file_paths, results):
        if result:
            await log_queue.put(f"Translated: {file_path} -> {result}")